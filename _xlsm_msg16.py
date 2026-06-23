import openpyxl
wb = openpyxl.load_workbook(
    r"C:\Users\dario.cescon\Dati\workspace-sts\Gas\Architettura\configuration_and_command_v1.21.xlsm",
    data_only=True, keep_vba=False)
ws = wb["822"]

print("=== H1 (raw payload) ===")
print(ws["H1"].value)

print("\n=== righe 187-260 colonna A (decoded ASCII) ===")
for r in range(187, 260):
    v = ws.cell(row=r, column=1).value
    if v is not None and v != "":
        print(f"A{r}: {v!r}")

print("\n=== righe 187-260 dettaglio multi-colonna per Msg#16 ===")
# Le righe 195-210 dovrebbero contenere i campi ASCII decodificati di Msg#16
for r in range(193, 215):
    row_vals = []
    for c in range(1, 9):
        v = ws.cell(row=r, column=c).value
        if v is not None and v != "":
            row_vals.append(f"{chr(64+c)}={v!r}")
    if row_vals:
        print(f"R{r:03d}: " + " | ".join(row_vals))
