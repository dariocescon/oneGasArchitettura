import openpyxl
wb = openpyxl.load_workbook(
    r"C:\Users\dario.cescon\Dati\workspace-sts\Gas\Architettura\configuration_and_command_v1.21.xlsm",
    data_only=True, keep_vba=False)
ws = wb["822"]

print("=== H1 (raw payload) ===")
print(ws["H1"].value)

print("\n=== righe 25-50 (header) colonna A+C+G+H ===")
for r in range(13, 50):
    row_vals = []
    for c in [1, 2, 3, 7, 8, 9]:
        v = ws.cell(row=r, column=c).value
        if v is not None and v != "":
            row_vals.append(f"{chr(64+c)}={v!r}")
    if row_vals:
        print(f"R{r:03d}: " + " | ".join(row_vals))

print("\n=== ricerca campi GPS (timeToFix, UTC, LAT, LON, ecc.) ===")
# I campi GPS dovrebbero comparire nella sezione di msg #17 dello sheet
# Magari intorno alle righe 30-50 visto che la struttura ASCII inizia
# subito dopo byte 16
for r in range(1, 250):
    h = ws.cell(row=r, column=8).value
    if h and isinstance(h, str) and any(k in h.lower() for k in
            ['gps', 'utc', 'lat', 'lon', 'altitude', 'speed', 'satellit', 'fix', 'cog', 'hdop', 'time to fix']):
        a = ws.cell(row=r, column=1).value
        c = ws.cell(row=r, column=3).value
        g = ws.cell(row=r, column=7).value
        print(f"R{r:03d}: A={a!r} | C={c!r} | G={g!r} | H={h!r}")
