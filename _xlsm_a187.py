import openpyxl
wb = openpyxl.load_workbook(
    r"C:\Users\dario.cescon\Dati\workspace-sts\Gas\Architettura\configuration_and_command_v1.21.xlsm",
    data_only=True, keep_vba=False)
ws = wb["822"]
print("H1 =", ws["H1"].value)
print("A187 =", ws["A187"].value)
# anche righe vicine per contesto
for r in range(180, 220):
    v = ws.cell(row=r, column=1).value
    if v is not None and v != "":
        print(f"A{r}: {v}")
