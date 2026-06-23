import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\dario.cescon\Dati\workspace-sts\Gas\Architettura\configuration_and_command_v1.21.xlsm", data_only=True, keep_vba=False)

def dump(sheet_name, out_path, max_col=20):
    ws = wb[sheet_name]
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"=== SHEET: {sheet_name} ({ws.max_row}x{ws.max_column}) ===\n")
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=min(ws.max_column, max_col), values_only=True), start=1):
            if all(c is None or c == "" for c in row):
                continue
            cells = [("" if c is None else str(c)) for c in row]
            f.write(f"R{r_idx:04d}: " + " | ".join(cells) + "\n")

dump("822", r"C:\Users\dario.cescon\Dati\workspace-sts\Gas\Architettura\_xlsm_822.txt", max_col=38)
dump("822 CC", r"C:\Users\dario.cescon\Dati\workspace-sts\Gas\Architettura\_xlsm_822_CC.txt", max_col=16)
dump("Request Commands", r"C:\Users\dario.cescon\Dati\workspace-sts\Gas\Architettura\_xlsm_RC.txt", max_col=7)
dump("Device info", r"C:\Users\dario.cescon\Dati\workspace-sts\Gas\Architettura\_xlsm_DI.txt", max_col=2)
dump("HMI", r"C:\Users\dario.cescon\Dati\workspace-sts\Gas\Architettura\_xlsm_HMI.txt", max_col=19)
dump("Dashboard", r"C:\Users\dario.cescon\Dati\workspace-sts\Gas\Architettura\_xlsm_DASH.txt", max_col=11)
print("ok")
